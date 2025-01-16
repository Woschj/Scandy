from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, g, send_file
from app.models.database import Database
from app.utils.decorators import admin_required, tech_required
from werkzeug.utils import secure_filename
import os
from flask import current_app
from app.utils.db_schema import SchemaManager
import colorsys
import logging
from datetime import datetime, timedelta
from app.models.models import Tool, Consumable, Worker
import sqlite3
from app.utils.error_handler import handle_errors, safe_db_query
from app.utils.color_settings import save_color_setting
from werkzeug.security import generate_password_hash
from app.models.settings import Settings
from app.config import Config
import openpyxl
from io import BytesIO
from pathlib import Path
from backup import DatabaseBackup

# Logger einrichten
logger = logging.getLogger(__name__)

bp = Blueprint('admin', __name__, url_prefix='/admin')
backup_manager = DatabaseBackup(app_path=Path(__file__).parent.parent.parent)

def get_recent_activity():
    """Hole die letzten Aktivitäten"""
    return Database.query('''
        SELECT 
            'Ausleihe' as type,
            t.name as item_name,
            w.firstname || ' ' || w.lastname as worker_name,
            l.lent_at as action_date
        FROM lendings l
        JOIN tools t ON l.tool_barcode = t.barcode
        JOIN workers w ON l.worker_barcode = w.barcode
        WHERE l.returned_at IS NULL
        LIMIT 6
        
        UNION ALL
        
        SELECT 
            'Verbrauch' as type,
            c.name as item_name,
            w.firstname || ' ' || w.lastname as worker_name,
            cu.used_at as action_date
        FROM consumable_usages cu
        JOIN consumables c ON cu.consumable_barcode = c.barcode
        JOIN workers w ON cu.worker_barcode = w.barcode
        ORDER BY action_date DESC
        LIMIT 10
    ''')

def get_material_usage():
    """Hole die Materialnutzung"""
    return Database.query('''
        SELECT 
            c.name,
            SUM(CASE WHEN cu.quantity > 0 THEN cu.quantity ELSE 0 END) as total_quantity
        FROM consumable_usages cu
        JOIN consumables c ON cu.consumable_barcode = c.barcode
        GROUP BY c.name
        ORDER BY total_quantity DESC
        LIMIT 5
    ''')

def get_warnings():
    """Hole aktuelle Warnungen"""
    return Database.query('''
        SELECT 
            'Werkzeug defekt' as type,
            name as message,
            'error' as severity
        FROM tools 
        WHERE status = 'defekt' AND deleted = 0
        
        UNION ALL
        
        SELECT 
            'Material niedrig' as type,
            name || ' (Bestand: ' || quantity || ')' as message,
            CASE 
                WHEN quantity < min_quantity * 0.5 THEN 'error'
                ELSE 'warning'
            END as severity
        FROM consumables
        WHERE quantity < min_quantity AND deleted = 0
        ORDER BY severity DESC
        LIMIT 5
    ''')

def get_backup_info():
    """Hole Informationen über vorhandene Backups"""
    backups = []
    backup_dir = Path(__file__).parent.parent.parent / 'backups'
    
    if backup_dir.exists():
        for backup_file in sorted(backup_dir.glob('*.db'), reverse=True):
            # Backup-Statistiken aus der Datei lesen
            stats = None
            try:
                with sqlite3.connect(str(backup_file)) as conn:
                    cursor = conn.cursor()
                    stats = {
                        'tools': cursor.execute('SELECT COUNT(*) FROM tools WHERE deleted = 0').fetchone()[0],
                        'consumables': cursor.execute('SELECT COUNT(*) FROM consumables WHERE deleted = 0').fetchone()[0],
                        'workers': cursor.execute('SELECT COUNT(*) FROM workers WHERE deleted = 0').fetchone()[0],
                        'active_lendings': cursor.execute('SELECT COUNT(*) FROM lendings WHERE returned_at IS NULL').fetchone()[0]
                    }
            except:
                stats = None
            
            # Unix-Timestamp in datetime umwandeln
            created_timestamp = backup_file.stat().st_mtime
            created_datetime = datetime.fromtimestamp(created_timestamp)
            
            backups.append({
                'name': backup_file.name,
                'size': backup_file.stat().st_size,
                'created': created_datetime,
                'stats': stats
            })
    
    return backups

def get_consumables_forecast():
    """Berechnet die Bestandsprognose für Verbrauchsmaterialien"""
    return Database.query('''
        WITH avg_usage AS (
            SELECT 
                c.barcode,
                c.name,
                c.quantity as current_amount,
                COALESCE(CAST(SUM(cu.quantity) AS FLOAT) / 30, 0) as avg_daily_usage
            FROM consumables c
            LEFT JOIN consumable_usages cu ON c.barcode = cu.consumable_barcode
                AND cu.used_at >= date('now', '-30 days')
            WHERE c.deleted = 0
            GROUP BY c.barcode, c.name, c.quantity
        )
        SELECT 
            name,
            current_amount,
            ROUND(avg_daily_usage, 2) as avg_daily_usage,
            CASE 
                WHEN avg_daily_usage > 0 THEN ROUND(current_amount / avg_daily_usage)
                ELSE 999
            END as days_remaining
        FROM avg_usage
        WHERE current_amount > 0
        ORDER BY 
            CASE 
                WHEN avg_daily_usage > 0 THEN current_amount / avg_daily_usage
                ELSE 999
            END ASC
        LIMIT 6
    ''')

def create_excel(data, columns):
    """Erstellt eine Excel-Datei aus den gegebenen Daten"""
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Headers hinzufügen
    for col, header in enumerate(columns, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Daten einfügen
    for row, item in enumerate(data, 2):
        for col, key in enumerate(columns, 1):
            ws.cell(row=row, column=col, value=item.get(key, ''))
    
    # Excel-Datei in BytesIO speichern
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return excel_file

@bp.route('/')
@admin_required
def index():
    return redirect(url_for('admin.dashboard'))

@bp.route('/dashboard')
@admin_required
def dashboard():
    """Admin Dashboard anzeigen"""
    stats = {
        'maintenance_issues': Database.query("""
            SELECT name, status, 
                CASE 
                    WHEN status = 'defekt' THEN 'error'
                    ELSE 'warning'
                END as severity
            FROM tools 
            WHERE status = 'defekt' AND deleted = 0
            ORDER BY name
        """),
        'inventory_warnings': Database.query("""
            SELECT
                name as message,
                CASE
                    WHEN quantity < min_quantity * 0.5 THEN 'error'
                    ELSE 'warning'
                END as type,
                CASE
                    WHEN quantity < min_quantity * 0.5 THEN 'exclamation-triangle'
                    ELSE 'exclamation'
                END as icon
            FROM consumables
            WHERE quantity < min_quantity AND deleted = 0
            ORDER BY quantity / min_quantity ASC
            LIMIT 5
        """),
        'consumable_trend': get_consumable_trend()
    }
    
    # Aktuelle Ausleihen laden
    current_lendings = Database.query("""
        SELECT l.*, t.name as tool_name, w.firstname || ' ' || w.lastname as worker_name
        FROM lendings l
        JOIN tools t ON l.tool_barcode = t.barcode
        JOIN workers w ON l.worker_barcode = w.barcode
        WHERE l.returned_at IS NULL
        ORDER BY l.lent_at DESC
    """)
    
    # Materialausgaben laden
    consumable_usages = Database.query("""
        SELECT 
            c.name as consumable_name,
            cu.quantity,
            w.firstname || ' ' || w.lastname as worker_name,
            cu.used_at
        FROM consumable_usages cu
        JOIN consumables c ON cu.consumable_barcode = c.barcode
        JOIN workers w ON cu.worker_barcode = w.barcode
        ORDER BY cu.used_at DESC
        LIMIT 10
    """)
    
    # Bestandsprognose laden
    consumables_forecast = get_consumables_forecast()
    
    # Backup-Informationen laden
    backups = get_backup_info()
    
    return render_template('admin/dashboard.html',
                         stats=stats,
                         current_lendings=current_lendings,
                         consumable_usages=consumable_usages,
                         consumables_forecast=consumables_forecast,
                         backups=backups,
                         Config=Config)

def get_consumable_trend():
    """Hole die Top 5 Materialverbrauch der letzten 7 Tage"""
    trend_data = Database.query('''
        WITH daily_usage AS (
            SELECT 
                c.name,
                date(cu.used_at) as date,
                SUM(CASE WHEN cu.quantity > 0 THEN cu.quantity ELSE 0 END) as daily_quantity
            FROM consumable_usages cu
            JOIN consumables c ON cu.consumable_barcode = c.barcode
            WHERE cu.used_at >= date('now', '-6 days')
            GROUP BY c.name, date(cu.used_at)
        ),
        dates AS (
            SELECT date('now', '-' || n || ' days') as date
            FROM (SELECT 0 as n UNION SELECT 1 UNION SELECT 2 UNION SELECT 3 
                 UNION SELECT 4 UNION SELECT 5 UNION SELECT 6)
        ),
        top_consumables AS (
            SELECT 
                c.name,
                SUM(CASE WHEN cu.quantity > 0 THEN cu.quantity ELSE 0 END) as total_quantity
            FROM consumable_usages cu
            JOIN consumables c ON cu.consumable_barcode = c.barcode
            WHERE cu.used_at >= date('now', '-6 days')
            GROUP BY c.name
            ORDER BY total_quantity DESC
            LIMIT 5
        )
        SELECT 
            dates.date as label,
            tc.name,
            COALESCE(du.daily_quantity, 0) as count
        FROM dates
        CROSS JOIN top_consumables tc
        LEFT JOIN daily_usage du ON dates.date = du.date AND tc.name = du.name
        ORDER BY tc.name, dates.date
    ''')
    
    # Daten für das Chart aufbereiten
    labels = []
    datasets = []
    current_name = None
    current_data = []
    
    for row in trend_data:
        if row['label'] not in labels:
            labels.append(row['label'])
        
        if current_name != row['name']:
            if current_name is not None:
                datasets.append({
                    'label': current_name,
                    'data': current_data,
                    'fill': False,
                    'borderColor': f'hsl({(len(datasets) * 60) % 360}, 70%, 50%)',
                    'tension': 0.1
                })
            current_name = row['name']
            current_data = []
        current_data.append(row['count'])
    
    if current_name is not None:
        datasets.append({
            'label': current_name,
            'data': current_data,
            'fill': False,
            'borderColor': f'hsl({(len(datasets) * 60) % 360}, 70%, 50%)',
            'tension': 0.1
        })
    
    return {
        'labels': labels,
        'datasets': datasets
    }

# Manuelle Ausleihe
@bp.route('/manual-lending', methods=['GET', 'POST'])
@admin_required
def manual_lending():
    """Manuelle Ausleihe/Rückgabe"""
    if request.method == 'POST':
        print("POST-Anfrage empfangen")
        
        try:
            # JSON-Daten verarbeiten
            data = request.get_json()
            if data is None:
                return jsonify({
                    'success': False,
                    'message': 'Keine gültigen JSON-Daten empfangen'
                }), 400
                
            print("JSON-Daten:", data)
            
            item_barcode = data.get('item_barcode')
            worker_barcode = data.get('worker_barcode')
            action = data.get('action')  # 'lend' oder 'return'
            quantity = data.get('quantity')
            if quantity is not None:
                quantity = int(quantity)
            
            if not item_barcode:
                return jsonify({
                    'success': False, 
                    'message': 'Artikel muss ausgewählt sein'
                }), 400
            
            if action == 'lend' and not worker_barcode:
                return jsonify({
                    'success': False, 
                    'message': 'Mitarbeiter muss ausgewählt sein'
                }), 400
            
            try:
                with Database.get_db() as db:
                    # Prüfe ob der Mitarbeiter existiert
                    if worker_barcode:
                        worker = db.execute('''
                            SELECT * FROM workers 
                            WHERE barcode = ? AND deleted = 0
                        ''', [worker_barcode]).fetchone()
                        
                        if not worker:
                            return jsonify({
                                'success': False,
                                'message': 'Mitarbeiter nicht gefunden'
                            }), 404
                    
                    # Prüfe ob es ein Verbrauchsmaterial ist
                    consumable = db.execute('''
                        SELECT * FROM consumables 
                        WHERE barcode = ? AND deleted = 0
                    ''', [item_barcode]).fetchone()
                    
                    if consumable:  # Verbrauchsmaterial-Logik
                        if not quantity or quantity <= 0:
                            return jsonify({
                                'success': False,
                                'message': 'Ungültige Menge'
                            }), 400
                            
                        if quantity > consumable['quantity']:
                            return jsonify({
                                'success': False,
                                'message': 'Nicht genügend Bestand'
                            }), 400
                            
                        # Neue Verbrauchsmaterial-Ausgabe erstellen
                        db.execute('''
                            INSERT INTO consumable_usages 
                            (consumable_barcode, worker_barcode, quantity, used_at, modified_at, sync_status)
                            VALUES (?, ?, ?, datetime('now'), datetime('now'), 'pending')
                        ''', [item_barcode, worker_barcode, quantity])
                        
                        # Bestand aktualisieren
                        db.execute('''
                            UPDATE consumables
                            SET quantity = quantity - ?,
                                modified_at = datetime('now'),
                                sync_status = 'pending'
                            WHERE barcode = ?
                        ''', [quantity, item_barcode])
                        
                        db.commit()
                        return jsonify({
                            'success': True,
                            'message': f'{quantity}x {consumable["name"]} wurde an {worker["firstname"]} {worker["lastname"]} ausgegeben'
                        })
                    
                    # Werkzeug-Logik
                    tool = db.execute('''
                        SELECT t.*, 
                               CASE 
                                   WHEN EXISTS (
                                       SELECT 1 FROM lendings l 
                                       WHERE l.tool_barcode = t.barcode 
                                       AND l.returned_at IS NULL
                                   ) THEN 'ausgeliehen'
                                   ELSE t.status
                               END as current_status,
                               (
                                   SELECT w.firstname || ' ' || w.lastname
                                   FROM lendings l
                                   JOIN workers w ON l.worker_barcode = w.barcode
                                   WHERE l.tool_barcode = t.barcode
                                   AND l.returned_at IS NULL
                                   LIMIT 1
                               ) as current_worker_name,
                               (
                                   SELECT l.worker_barcode
                                   FROM lendings l
                                   WHERE l.tool_barcode = t.barcode
                                   AND l.returned_at IS NULL
                                   LIMIT 1
                               ) as current_worker_barcode
                        FROM tools t
                        WHERE t.barcode = ? AND t.deleted = 0
                    ''', [item_barcode]).fetchone()
                    
                    if not tool:
                        return jsonify({
                            'success': False,
                            'message': 'Werkzeug nicht gefunden'
                        }), 404
                    
                    if action == 'lend':
                        if tool['current_status'] == 'ausgeliehen':
                            return jsonify({
                                'success': False,
                                'message': f'Dieses Werkzeug ist bereits an {tool["current_worker_name"]} ausgeliehen'
                            }), 400
                            
                        if tool['current_status'] == 'defekt':
                            return jsonify({
                                'success': False,
                                'message': 'Dieses Werkzeug ist als defekt markiert'
                            }), 400
                        
                        # Neue Ausleihe erstellen
                        db.execute('''
                            INSERT INTO lendings 
                            (tool_barcode, worker_barcode, lent_at, modified_at, sync_status)
                            VALUES (?, ?, datetime('now'), datetime('now'), 'pending')
                        ''', [item_barcode, worker_barcode])
                        
                        # Status des Werkzeugs aktualisieren
                        db.execute('''
                            UPDATE tools 
                            SET status = 'ausgeliehen',
                                modified_at = datetime('now'),
                                sync_status = 'pending'
                            WHERE barcode = ?
                        ''', [item_barcode])
                        
                        db.commit()
                        return jsonify({
                            'success': True,
                            'message': f'Werkzeug {tool["name"]} wurde an {worker["firstname"]} {worker["lastname"]} ausgeliehen'
                        })
                        
                    else:  # action == 'return'
                        if tool['current_status'] != 'ausgeliehen':
                            return jsonify({
                                'success': False,
                                'message': 'Dieses Werkzeug ist nicht ausgeliehen'
                            }), 400
                        
                        # Wenn ein Mitarbeiter angegeben wurde, prüfe ob er berechtigt ist
                        if worker_barcode and tool['current_worker_barcode'] != worker_barcode:
                            return jsonify({
                                'success': False,
                                'message': f'Dieses Werkzeug wurde von {tool["current_worker_name"]} ausgeliehen'
                            }), 403
                        
                        # Rückgabe verarbeiten
                        db.execute('''
                            UPDATE lendings 
                            SET returned_at = datetime('now'),
                                modified_at = datetime('now'),
                                sync_status = 'pending'
                            WHERE tool_barcode = ? 
                            AND returned_at IS NULL
                        ''', [item_barcode])
                        
                        # Status des Werkzeugs aktualisieren
                        db.execute('''
                            UPDATE tools 
                            SET status = 'verfügbar',
                                modified_at = datetime('now'),
                                sync_status = 'pending'
                            WHERE barcode = ?
                        ''', [item_barcode])
                        
                        db.commit()
                        return jsonify({
                            'success': True,
                            'message': f'Werkzeug {tool["name"]} wurde zurückgegeben'
                        })
                    
            except Exception as e:
                print("Fehler bei der Ausleihe:", str(e))
                return jsonify({
                    'success': False, 
                    'message': f'Fehler: {str(e)}'
                }), 500
            
        except Exception as e:
            print(f"Fehler beim Verarbeiten der Anfrage: {str(e)}")
            return jsonify({
                'success': False,
                'message': 'Fehler beim Verarbeiten der Anfrage'
            }), 500
            
    # GET request - zeige das Formular
    tools = Database.query('''
        SELECT t.*,
               CASE 
                   WHEN EXISTS (
                       SELECT 1 FROM lendings l 
                       WHERE l.tool_barcode = t.barcode 
                       AND l.returned_at IS NULL
                   ) THEN 'ausgeliehen'
                   WHEN t.status = 'defekt' THEN 'defekt'
                   ELSE 'verfügbar'
               END as current_status,
               (
                   SELECT w.firstname || ' ' || w.lastname
                   FROM lendings l
                   JOIN workers w ON l.worker_barcode = w.barcode
                   WHERE l.tool_barcode = t.barcode
                   AND l.returned_at IS NULL
                   LIMIT 1
               ) as current_worker,
               (
                   SELECT l.lent_at
                   FROM lendings l
                   WHERE l.tool_barcode = t.barcode
                   AND l.returned_at IS NULL
                   LIMIT 1
               ) as lending_date
        FROM tools t
        WHERE t.deleted = 0
        ORDER BY t.name
    ''')

    workers = Database.query('''
        SELECT * FROM workers WHERE deleted = 0
        ORDER BY lastname, firstname
    ''')

    consumables = Database.query('''
        SELECT c.*,
               CASE 
                   WHEN quantity <= 0 THEN 'nicht_verfügbar'
                   WHEN quantity <= min_quantity THEN 'kritisch'
                   WHEN quantity <= min_quantity * 2 THEN 'niedrig'
                   ELSE 'verfügbar'
               END as status
        FROM consumables c 
        WHERE deleted = 0
        ORDER BY name
    ''')

    # Aktuelle Ausleihen (Werkzeuge und Verbrauchsmaterial)
    current_lendings = Database.query('''
        WITH RECURSIVE current_tool_lendings AS (
            SELECT 
                l.id,
                t.name as item_name,
                t.barcode as item_barcode,
                w.firstname || ' ' || w.lastname as worker_name,
                w.barcode as worker_barcode,
                'Werkzeug' as category,
                l.lent_at as action_date,
                NULL as amount
            FROM lendings l
            JOIN tools t ON l.tool_barcode = t.barcode
            JOIN workers w ON l.worker_barcode = w.barcode
            WHERE l.returned_at IS NULL
            AND t.deleted = 0
            AND w.deleted = 0
        ),
        current_consumable_usages AS (
            SELECT 
                cu.id,
                c.name as item_name,
                c.barcode as item_barcode,
                w.firstname || ' ' || w.lastname as worker_name,
                w.barcode as worker_barcode,
                'Verbrauchsmaterial' as category,
                cu.used_at as action_date,
                cu.quantity as amount
            FROM consumable_usages cu
            JOIN consumables c ON cu.consumable_barcode = c.barcode
            JOIN workers w ON cu.worker_barcode = w.barcode
            WHERE DATE(cu.used_at) >= DATE('now', '-7 days')
            AND c.deleted = 0
            AND w.deleted = 0
        )
        SELECT * FROM current_tool_lendings
        UNION ALL
        SELECT * FROM current_consumable_usages
        ORDER BY action_date DESC
    ''')

    return render_template('admin/manual_lending.html', 
                          tools=tools,
                          workers=workers,
                          consumables=consumables,
                          current_lendings=current_lendings)

@bp.route('/trash')
@admin_required
def trash():
    """Papierkorb mit gelöschten Einträgen"""
    tools = Database.query('''
        SELECT * FROM tools WHERE deleted = 1
    ''')
    consumables = Database.query('''
        SELECT * FROM consumables WHERE deleted = 1
    ''')
    workers = Database.query('''
        SELECT * FROM workers WHERE deleted = 1
    ''')
    
    return render_template('admin/trash.html',
                         tools=tools,
                         consumables=consumables,
                         workers=workers)

@bp.route('/tools/<barcode>/delete', methods=['DELETE'])
@admin_required
def delete_tool(barcode):
    """Werkzeug in den Papierkorb verschieben"""
    try:
        Database.query('''
            UPDATE tools 
            SET deleted = 1,
                deleted_at = datetime('now'),
                modified_at = datetime('now'),
                sync_status = 'pending'
            WHERE barcode = ?
        ''', [barcode])
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@bp.route('/consumables/<barcode>/delete', methods=['DELETE'])
@admin_required
def delete_consumable(barcode):
    """Verbrauchsmaterial in den Papierkorb verschieben"""
    try:
        Database.query('''
            UPDATE consumables 
            SET deleted = 1,
                deleted_at = datetime('now'),
                modified_at = datetime('now'),
                sync_status = 'pending'
            WHERE barcode = ?
        ''', [barcode])
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@bp.route('/workers/<barcode>/delete', methods=['DELETE'])
@admin_required
def delete_worker(barcode):
    """Mitarbeiter in den Papierkorb verschieben"""
    try:
        Database.query('''
            UPDATE workers 
            SET deleted = 1,
                deleted_at = datetime('now'),
                modified_at = datetime('now'),
                sync_status = 'pending'
            WHERE barcode = ?
        ''', [barcode])
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@bp.route('/trash/restore/<type>/<barcode>', methods=['POST'])
@admin_required
def restore_item(type, barcode):
    """Element aus dem Papierkorb wiederherstellen"""
    try:
        table = {
            'tool': 'tools',
            'consumable': 'consumables',
            'worker': 'workers'
        }.get(type)
        
        if not table:
            return jsonify({'success': False, 'error': 'Ungültiger Typ'}), 400
            
        Database.query(f'''
            UPDATE {table}
            SET deleted = 0,
                deleted_at = NULL,
                modified_at = datetime('now'),
                sync_status = 'pending'
            WHERE barcode = ?
        ''', [barcode])
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@bp.route('/server-settings', methods=['GET', 'POST'])
@admin_required
def server_settings():
    if request.method == 'POST':
        mode = request.form.get('mode')
        server_url = request.form.get('server_url')
        
        try:
            # Speichere Einstellungen in der Datenbank
            Database.query('''
                INSERT OR REPLACE INTO settings (key, value)
                VALUES (?, ?)
            ''', ['server_mode', '1' if mode == 'server' else '0'])
            
            if mode == 'client' and server_url:
                Database.query('''
                    INSERT OR REPLACE INTO settings (key, value)
                    VALUES (?, ?)
                ''', ['server_url', server_url])
            
            if mode == 'server':
                Config.init_server()
                flash('Server-Modus aktiviert', 'success')
            else:
                Config.init_client(server_url)
                flash('Client-Modus aktiviert', 'success')
                
        except Exception as e:
            flash(f'Fehler beim Speichern der Einstellungen: {str(e)}', 'error')
            return redirect(url_for('admin.server_settings'))
            
    try:
        # Hole aktuelle Einstellungen
        status = Database.query('''
            SELECT last_sync, status, error 
            FROM sync_status 
            ORDER BY id DESC LIMIT 1
        ''', one=True)
        
        auto_sync = Database.query('''
            SELECT value FROM settings
            WHERE key = 'auto_sync'
        ''', one=True)
                
        return render_template('admin/server_settings.html',
                             is_server=Config.SERVER_MODE,
                             server_url=Config.SERVER_URL,
                             last_sync=status['last_sync'] if status else None,
                             sync_status=status['status'] if status else 'never',
                             sync_error=status['error'] if status else None,
                             auto_sync=bool(int(auto_sync['value'])) if auto_sync else False)
                             
    except Exception as e:
        flash(f'Fehler beim Laden der Einstellungen: {str(e)}', 'error')
        return redirect(url_for('admin.dashboard'))

@bp.route('/export/<table>')
@admin_required
def export_table(table):
    """Exportiert eine Tabelle als Excel"""
    try:
        with Database.get_db() as conn:
            if table == 'tools':
                query = """
                    SELECT 
                        t.name, t.barcode, t.category, t.location,
                        CASE WHEN l.id IS NULL THEN 'Nein' ELSE 'Ja' END as is_lent
                    FROM tools t
                    LEFT JOIN lendings l ON t.barcode = l.tool_barcode AND l.returned_at IS NULL
                    WHERE t.deleted = 0
                """
                rows = [dict(row) for row in conn.execute(query).fetchall()]
                excel_file = create_excel(rows, ['name', 'barcode', 'category', 'location', 'is_lent'])
                filename = f'werkzeuge_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
                
            elif table == 'workers':
                query = """
                    SELECT 
                        w.firstname, w.lastname, w.department, w.barcode,
                        COUNT(DISTINCT l.id) as active_lendings
                    FROM workers w
                    LEFT JOIN lendings l ON w.barcode = l.worker_barcode AND l.returned_at IS NULL
                    WHERE w.deleted = 0
                    GROUP BY w.id
                """
                rows = [dict(row) for row in conn.execute(query).fetchall()]
                excel_file = create_excel(rows, ['firstname', 'lastname', 'department', 'barcode', 'active_lendings'])
                filename = f'mitarbeiter_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
                
            elif table == 'consumables':
                query = """
                    SELECT name, barcode, category, location, quantity, min_quantity
                    FROM consumables 
                    WHERE deleted = 0
                """
                rows = [dict(row) for row in conn.execute(query).fetchall()]
                excel_file = create_excel(rows, ['name', 'barcode', 'category', 'location', 'quantity', 'min_quantity'])
                filename = f'verbrauchsmaterial_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
                
            elif table == 'history':
                query = """
                    SELECT 
                        l.lent_at, l.returned_at, t.name as item_name, 
                        w.firstname || ' ' || w.lastname as worker_name,
                        'Werkzeug' as type
                    FROM lendings l
                    JOIN tools t ON l.tool_barcode = t.barcode
                    JOIN workers w ON l.worker_barcode = w.barcode
                    UNION ALL
                    SELECT 
                        cu.used_at as lent_at, 
                        NULL as returned_at, 
                        c.name as item_name,
                        w.firstname || ' ' || w.lastname as worker_name,
                        'Verbrauchsmaterial' as type
                    FROM consumable_usages cu
                    JOIN consumables c ON cu.consumable_barcode = c.barcode
                    JOIN workers w ON cu.worker_barcode = w.barcode
                    ORDER BY lent_at DESC
                """
                rows = [dict(row) for row in conn.execute(query).fetchall()]
                excel_file = create_excel(rows, ['lent_at', 'returned_at', 'item_name', 'worker_name', 'type'])
                filename = f'verlauf_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            
            else:
                return jsonify({'success': False, 'message': 'Ungültige Tabelle'}), 400
            
            return send_file(
                excel_file,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=filename
            )
            
    except Exception as e:
        print(f"Export-Fehler: {str(e)}")
        return redirect(url_for('admin.dashboard'))

@bp.route('/import/<table>', methods=['POST'])
@admin_required
def import_table(table):
    """Importiert eine Excel-Tabelle"""
    if 'file' not in request.files:
        flash('Keine Datei ausgewählt', 'error')
        return redirect(url_for('admin.dashboard'))
        
    file = request.files['file']
    if file.filename == '':
        flash('Keine Datei ausgewählt', 'error')
        return redirect(url_for('admin.dashboard'))

    try:
        # Excel-Datei in BytesIO laden
        file_content = BytesIO(file.read())
        wb = openpyxl.load_workbook(file_content)
        ws = wb.active
        
        # Headers aus erster Zeile lesen
        headers = [cell.value for cell in ws[1]]
        
        if table == 'tools':
            for row in ws.iter_rows(min_row=2):
                row_data = {headers[i]: cell.value for i, cell in enumerate(row)}
                Database.query('''
                    INSERT OR REPLACE INTO tools 
                    (barcode, name, description, category, location)
                    VALUES (?, ?, ?, ?, ?)
                ''', [row_data.get('barcode'), row_data.get('name'), 
                     row_data.get('description'), row_data.get('category'), 
                     row_data.get('location')])
                
        elif table == 'workers':
            for row in ws.iter_rows(min_row=2):
                row_data = {headers[i]: cell.value for i, cell in enumerate(row)}
                Database.query('''
                    INSERT OR REPLACE INTO workers 
                    (barcode, firstname, lastname, department, email)
                    VALUES (?, ?, ?, ?, ?)
                ''', [row_data.get('barcode'), row_data.get('firstname'), row_data.get('lastname'),
                     row_data.get('department'), row_data.get('email')])
                
        elif table == 'consumables':
            for row in ws.iter_rows(min_row=2):
                row_data = {headers[i]: cell.value for i, cell in enumerate(row)}
                Database.query('''
                    INSERT OR REPLACE INTO consumables 
                    (barcode, name, description, quantity, min_quantity, category, location)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                ''', [row_data.get('barcode'), row_data.get('name'), row_data.get('description'),
                     row_data.get('quantity'), row_data.get('min_quantity'),
                     row_data.get('category'), row_data.get('location')])
                
        flash(f'Import erfolgreich', 'success')
        
    except Exception as e:
        flash(f'Fehler beim Import: {str(e)}', 'error')
        
    return redirect(url_for('admin.dashboard'))

@bp.route('/backup/create', methods=['POST'])
@admin_required
def create_backup():
    """Manuelles Backup erstellen"""
    try:
        logger.info("Starte manuelles Backup...")
        success = backup_manager.create_backup()
        if success:
            flash('Backup wurde erfolgreich erstellt', 'success')
            logger.info("Backup erfolgreich erstellt")
        else:
            flash('Fehler beim Erstellen des Backups', 'error')
            logger.error("Backup konnte nicht erstellt werden")
    except Exception as e:
        error_msg = f'Fehler beim Erstellen des Backups: {str(e)}'
        logger.error(error_msg)
        flash(error_msg, 'error')
    return redirect(url_for('admin.dashboard'))

@bp.route('/backup/restore/<filename>', methods=['POST'])
@admin_required
def restore_backup(filename):
    """Backup wiederherstellen"""
    try:
        logger.info(f"Versuche Backup wiederherzustellen: {filename}")
        backup_path = backup_manager.backup_dir / filename
        if not backup_path.exists():
            error_msg = 'Backup-Datei nicht gefunden'
            logger.error(f"{error_msg}: {backup_path}")
            flash(error_msg, 'error')
            return redirect(url_for('admin.dashboard'))
        
        # Datenbank wiederherstellen
        success = backup_manager.restore_backup(filename)
        if success:
            success_msg = 'Backup wurde erfolgreich wiederhergestellt'
            logger.info(success_msg)
            flash(success_msg, 'success')
        else:
            error_msg = 'Fehler bei der Wiederherstellung des Backups'
            logger.error(error_msg)
            flash(error_msg, 'error')
    except Exception as e:
        error_msg = f'Fehler beim Wiederherstellen des Backups: {str(e)}'
        logger.error(error_msg)
        flash(error_msg, 'error')
    return redirect(url_for('admin.dashboard'))

@bp.route('/departments', methods=['GET'])
@admin_required
def get_departments():
    """Hole alle Abteilungen"""
    try:
        departments = Database.query('''
            SELECT value as name
            FROM settings
            WHERE key LIKE 'department_%'
            AND value IS NOT NULL
            AND value != ''
            ORDER BY value
        ''')
        return jsonify({
            'success': True,
            'departments': [dept['name'] for dept in departments]
        })
    except Exception as e:
        print(f"Fehler beim Laden der Abteilungen: {e}")
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500

@bp.route('/departments/add', methods=['POST'])
@admin_required
def add_department():
    """Füge eine neue Abteilung hinzu"""
    data = request.get_json()
    name = data.get('name')
    if not name:
        return jsonify({'success': False, 'message': 'Kein Name angegeben'})
    
    try:
        db = Database.get_db()
        # Nächste freie ID finden
        result = db.execute("""
            SELECT MAX(CAST(SUBSTR(key, 12) AS INTEGER)) as max_id 
            FROM settings 
            WHERE key LIKE 'department_%'
        """).fetchone()
        next_id = 1 if not result or result['max_id'] is None else result['max_id'] + 1
        
        # Neue Abteilung einfügen
        db.execute(
            "INSERT INTO settings (key, value) VALUES (?, ?)",
            [f'department_{next_id}', name]
        )
        db.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@bp.route('/departments/<name>', methods=['DELETE'])
@admin_required
def delete_department(name):
    """Lösche eine Abteilung"""
    try:
        Database.execute(
            "DELETE FROM settings WHERE key LIKE 'department_%' AND value = ?",
            [name]
        )
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

# Standortverwaltung
@bp.route('/locations', methods=['GET'])
@admin_required
def get_locations():
    """Hole alle Standorte"""
    try:
        locations = Database.query('''
            SELECT value as name, description
            FROM settings
            WHERE key LIKE 'location_%'
            AND value IS NOT NULL
            AND value != ''
            ORDER BY value
        ''')
        return jsonify({
            'success': True,
            'locations': locations
        })
    except Exception as e:
        print(f"Fehler beim Laden der Standorte: {e}")
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500

@bp.route('/locations/add', methods=['POST'])
@admin_required
def add_location():
    """Füge einen neuen Standort hinzu"""
    data = request.get_json()
    name = data.get('name')
    if not name:
        return jsonify({'success': False, 'message': 'Kein Name angegeben'})
    
    try:
        db = Database.get_db()
        # Nächste freie ID finden
        result = db.execute("""
            SELECT MAX(CAST(SUBSTR(key, 10) AS INTEGER)) as max_id 
            FROM settings 
            WHERE key LIKE 'location_%'
        """).fetchone()
        next_id = 1 if not result or result['max_id'] is None else result['max_id'] + 1
        
        # Neuen Standort einfügen
        db.execute(
            "INSERT INTO settings (key, value, description) VALUES (?, ?, ?)",
            [f'location_{next_id}', name, 'both']
        )
        db.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@bp.route('/locations/<name>', methods=['DELETE'])
@admin_required
def delete_location(name):
    """Lösche einen Standort"""
    try:
        Database.execute(
            "DELETE FROM settings WHERE key LIKE 'location_%' AND value = ?",
            [name]
        )
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

# Kategorieverwaltung
@bp.route('/categories', methods=['GET'])
@admin_required
def get_categories():
    """Hole alle Kategorien"""
    try:
        categories = Database.query('''
            SELECT value as name, description
            FROM settings
            WHERE key LIKE 'category_%'
            AND value IS NOT NULL
            AND value != ''
            ORDER BY value
        ''')
        return jsonify({
            'success': True,
            'categories': categories
        })
    except Exception as e:
        print(f"Fehler beim Laden der Kategorien: {e}")
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500

@bp.route('/categories/add', methods=['POST'])
@admin_required
def add_category():
    """Füge eine neue Kategorie hinzu"""
    data = request.get_json()
    name = data.get('name')
    if not name:
        return jsonify({'success': False, 'message': 'Kein Name angegeben'})
    
    try:
        db = Database.get_db()
        # Nächste freie ID finden
        result = db.execute("""
            SELECT MAX(CAST(SUBSTR(key, 10) AS INTEGER)) as max_id 
            FROM settings 
            WHERE key LIKE 'category_%'
        """).fetchone()
        next_id = 1 if not result or result['max_id'] is None else result['max_id'] + 1
        
        # Neue Kategorie einfügen
        db.execute(
            "INSERT INTO settings (key, value) VALUES (?, ?)",
            [f'category_{next_id}', name]
        )
        db.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@bp.route('/categories/<name>', methods=['DELETE'])
@admin_required
def delete_category(name):
    """Lösche eine Kategorie"""
    try:
        Database.execute(
            "DELETE FROM settings WHERE key LIKE 'category_%' AND value = ?",
            [name]
        )
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@bp.route('/locations/<name>', methods=['PUT'])
@admin_required
def update_location(name):
    """Aktualisiere einen Standort"""
    location = request.form.get('location')
    tools = request.form.get('tools') == 'true'
    consumables = request.form.get('consumables') == 'true'
    
    if not location:
        return jsonify({
            'success': False,
            'message': 'Standortname fehlt'
        }), 400

    try:
        # Finde den Basis-Schlüssel für den Standort
        location_key = Database.query('''
            SELECT key FROM settings 
            WHERE key LIKE 'location_%'
            AND key NOT LIKE '%_tools'
            AND key NOT LIKE '%_consumables'
            AND value = ?
        ''', [name], one=True)
        
        if not location_key:
            return jsonify({
                'success': False,
                'message': 'Standort nicht gefunden'
            }), 404
            
        base_key = location_key['key']
        
        # Aktualisiere den Standort und seine Eigenschaften
        Database.query('''
            UPDATE settings 
            SET value = ?, modified_at = datetime('now'), sync_status = 'pending'
            WHERE key = ?
        ''', [location, base_key])
        
        Database.query('''
            UPDATE settings 
            SET value = ?, modified_at = datetime('now'), sync_status = 'pending'
            WHERE key = ?
        ''', ['1' if tools else '0', f'{base_key}_tools'])
        
        Database.query('''
            UPDATE settings 
            SET value = ?, modified_at = datetime('now'), sync_status = 'pending'
            WHERE key = ?
        ''', ['1' if consumables else '0', f'{base_key}_consumables'])

        return jsonify({
            'success': True,
            'message': 'Standort erfolgreich aktualisiert'
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500

@bp.route('/categories/<name>', methods=['PUT'])
@admin_required
def update_category(name):
    """Aktualisiere eine Kategorie"""
    category = request.form.get('category')
    tools = request.form.get('tools') == 'true'
    consumables = request.form.get('consumables') == 'true'
    
    if not category:
        return jsonify({
            'success': False,
            'message': 'Kategoriename fehlt'
        }), 400

    try:
        # Finde den Basis-Schlüssel für die Kategorie
        category_key = Database.query('''
            SELECT key FROM settings 
            WHERE key LIKE 'category_%'
            AND key NOT LIKE '%_tools'
            AND key NOT LIKE '%_consumables'
            AND value = ?
        ''', [name], one=True)
        
        if not category_key:
            return jsonify({
                'success': False,
                'message': 'Kategorie nicht gefunden'
            }), 404
            
        base_key = category_key['key']
        
        # Aktualisiere die Kategorie und ihre Eigenschaften
        Database.query('''
            UPDATE settings 
            SET value = ?, modified_at = datetime('now'), sync_status = 'pending'
            WHERE key = ?
        ''', [category, base_key])
        
        Database.query('''
            UPDATE settings 
            SET value = ?, modified_at = datetime('now'), sync_status = 'pending'
            WHERE key = ?
        ''', ['1' if tools else '0', f'{base_key}_tools'])
        
        Database.query('''
            UPDATE settings 
            SET value = ?, modified_at = datetime('now'), sync_status = 'pending'
            WHERE key = ?
        ''', ['1' if consumables else '0', f'{base_key}_consumables'])

        return jsonify({
            'success': True,
            'message': 'Kategorie erfolgreich aktualisiert'
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500

@bp.route('/cleanup-database', methods=['POST'])
@admin_required
def cleanup_database():
    """Bereinigt die Datenbank von ungültigen Einträgen"""
    try:
        print("Starte Datenbankbereinigung...")
        with Database.get_db() as db:
            # Finde Ausleihungen für nicht existierende Werkzeuge
            invalid_lendings = db.execute('''
                SELECT l.*, t.barcode as tool_exists
                FROM lendings l
                LEFT JOIN tools t ON l.tool_barcode = t.barcode
                WHERE t.barcode IS NULL
                AND l.returned_at IS NULL
            ''').fetchall()
            
            print(f"Gefundene ungültige Ausleihungen: {len(invalid_lendings)}")
            for lending in invalid_lendings:
                print(f"Ungültige Ausleihe gefunden: {dict(lending)}")
            
            # Lösche ungültige Ausleihungen
            if invalid_lendings:
                db.execute('''
                    DELETE FROM lendings
                    WHERE id IN (
                        SELECT l.id
                        FROM lendings l
                        LEFT JOIN tools t ON l.tool_barcode = t.barcode
                        WHERE t.barcode IS NULL
                        AND l.returned_at IS NULL
                    )
                ''')
                db.commit()
                print(f"{len(invalid_lendings)} Ausleihungen wurden gelöscht")
                
            return jsonify({
                'success': True,
                'message': f'{len(invalid_lendings)} ungültige Ausleihungen wurden bereinigt'
            })
            
    except Exception as e:
        print(f"Fehler bei der Bereinigung: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Fehler bei der Bereinigung: {str(e)}'
        }), 500

@bp.route('/departments/list')
@admin_required
def list_departments():
    """Liste alle Abteilungen auf"""
    departments = Database.query("""
        SELECT value as name 
        FROM settings 
        WHERE key LIKE 'department_%'
        AND value IS NOT NULL 
        AND value != ''
        ORDER BY value
    """)
    return jsonify({'success': True, 'departments': [dict(dept) for dept in departments]})

@bp.route('/locations/list')
@admin_required
def list_locations():
    """Liste alle Standorte auf"""
    locations = Database.query("""
        SELECT value as name, description as usage
        FROM settings 
        WHERE key LIKE 'location_%'
        AND value IS NOT NULL 
        AND value != ''
        ORDER BY value
    """)
    return jsonify({'success': True, 'locations': [dict(loc) for loc in locations]})

@bp.route('/categories/list')
@admin_required
def list_categories():
    """Liste alle Kategorien auf"""
    categories = Database.query("""
        SELECT value as name, description as usage
        FROM settings 
        WHERE key LIKE 'category_%'
        AND value IS NOT NULL 
        AND value != ''
        ORDER BY value
    """)
    return jsonify({'success': True, 'categories': [dict(cat) for cat in categories]})