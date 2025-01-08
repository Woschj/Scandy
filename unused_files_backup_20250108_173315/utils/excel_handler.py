import sqlite3
import os
from datetime import datetime
from openpyxl import Workbook

class ExcelHandler:
    @staticmethod
    def export_to_excel(db_path, output_dir='exports'):
        try:
            os.makedirs(output_dir, exist_ok=True)
            conn = sqlite3.connect(db_path)
            cursor = conn.cursor()
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
            tables = cursor.fetchall()
            
            db_name = os.path.splitext(os.path.basename(db_path))[0]
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            excel_path = os.path.join(output_dir, f'{db_name}_{timestamp}.xlsx')
            
            wb = Workbook()
            
            for table in tables:
                table_name = table[0]
                cursor.execute(f"SELECT * FROM {table_name}")
                columns = [description[0] for description in cursor.description]
                rows = cursor.fetchall()
                
                # Erstelle ein neues Worksheet für jede Tabelle
                if table == tables[0]:
                    ws = wb.active
                    ws.title = table_name
                else:
                    ws = wb.create_sheet(title=table_name)
                
                # Schreibe die Spaltenüberschriften
                for col, header in enumerate(columns, 1):
                    ws.cell(row=1, column=col, value=header)
                
                # Schreibe die Daten
                for row_idx, row in enumerate(rows, 2):
                    for col_idx, value in enumerate(row, 1):
                        ws.cell(row=row_idx, column=col_idx, value=value)
            
            wb.save(excel_path)
            return True, excel_path
            
        except Exception as e:
            return False, str(e) 